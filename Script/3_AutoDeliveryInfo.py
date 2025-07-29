import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
import time
from collections import defaultdict
import sys
from io import StringIO
import datetime


Path_SKU_KG = '/Users/lifenew/WorkSpace/Python/ExcelScrips/data/0715/产品预报信息表更新0712.xlsx'
Path_SKU_KG_Sheet = '产品预报明细表更新'

Path_Stores = '/Users/lifenew/WorkSpace/Python/ExcelScrips/data/0715/装箱单/A/A1'
Ware_House_Header = '"SKU","商品名称"'

#1.  产品预报信息结构
# Excel D E F G
# sku -> [0]净重kg [1]毛重kg [2]体积m³
SKU_KG = defaultdict(list)
SKU  = 'B'
SKU_N = 'C'
P_KG = 'E'
N_KG = 'F'
M3   = 'G'

def Read_SKU_KG_Info():
    global Path_SKU_KG
    try:
        # 打开工作簿
        workbook = openpyxl.load_workbook(Path_SKU_KG, data_only=True, read_only=True)
        # 获取所有表名
        sheet_names = workbook.sheetnames

        if Path_SKU_KG_Sheet in sheet_names:
            sheet = workbook[Path_SKU_KG_Sheet]
            
            print(sheet.max_row + 1)
            for idx in range(2, sheet.max_row + 1):
                sku_kg_data_idx = [ord(SKU) - ord('A'), ord(SKU_N) - ord('A'), ord(P_KG) - ord('A'), ord(N_KG) - ord('A'), ord(M3) - ord('A')]
                sku_kg_data = [sheet.cell(row=idx, column=i+1).value for i in sku_kg_data_idx]
                
                if all(cell is None or str(cell).strip() == '' for cell in sku_kg_data):
                    continue
                SKU_KG[sku_kg_data[0]] = [sku_kg_data[1], float(sku_kg_data[2]), float(sku_kg_data[3]), sku_kg_data[4]]
                # print(sku_kg_data[0], str(SKU_KG[sku_kg_data[0]]))
        else:
            print(f"Error: Sheet '{Path_SKU_KG_Sheet}' not exits!")
        
    except FileNotFoundError:
        print(f"Error: File '{Path_SKU_KG}' not exits!")
    except Exception as e:
        print(f"Error: {e}")
    return None

def GeneratorExcelDetail(warehouse_dict_lists_detail):

    def generate_random_data(warehouse_dict_lists_detail):
        data = []
        for location_k, location_v in warehouse_dict_lists_detail.items():
            for sku_k, sku_v in location_v.items():
                # 添加到数据
                data.append({
                    '出货日期': "0X3F",
                    '货件单号': location_k.split('-')[0],
                    'Reference ID': '',
                    '物流中心编码': location_k.split('-')[1],
                    '序列号': sku_v[3],
                    'SKU': sku_k,
                    'FNSKU': sku_v[1],
                    '品名': sku_v[2],
                    '发货数量': sku_v[0],
                })
        return data
    
    data = generate_random_data(warehouse_dict_lists_detail)

    print('Start')
    # 创建工作簿
    # 样式定义
    center_alignment = Alignment(horizontal='center', vertical='center')
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    # 创建工作表
    ws = wb.create_sheet(title='发货数据汇总')
    
    headers = [
        "出货日期", "货件单号", "Reference ID", "物流中心编码", "序列号", "SKU", "FNSKU", "品名",
        "发货数量", "同尺寸数量", "合计数量", "合计净重kg", "同尺寸净重", "净重合计",
        "单件毛重KG", "合计毛重kg", "同尺寸毛重", "毛重kg合计", "合计体积m³", "同尺寸体积", "合计体积",
        "尺寸", "货代", "渠道", "货代运单号", "备注", "货代收货地址"
    ]

    # 写入表头（加粗+居中）
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = center_alignment

    # 逐行写入数据（没有指定的key自动留空）
    for row_num, row_dict in enumerate(data, 2):  # 从第2行开始
        for col_num, header in enumerate(headers, 1):
            if header in row_dict:  # 如果该列在字典中有定义，则写入值
                ws.cell(row=row_num, column=col_num, value=row_dict[header])
            # 否则留空（无需操作）

    # 4. 自动合并相同内容的列（关键部分）
    merge_columns = ["出货日期", "货件单号", "物流中心编码"]  # 需要合并的列名

    for col_name in merge_columns:
        col_idx = headers.index(col_name) + 1  # 获取列索引（从1开始）
        start_row = 2
        current_value = ws.cell(row=start_row, column=col_idx).value
        
        for row_num in range(3, len(data) + 2):  # 从第3行开始检查
            cell_value = ws.cell(row=row_num, column=col_idx).value
            if cell_value != current_value:
                # 合并之前的相同单元格
                if start_row < row_num - 1:
                    ws.merge_cells(
                        start_row=start_row, end_row=row_num - 1,
                        start_column=col_idx, end_column=col_idx
                    )
                    # 合并后重新设置居中
                    for row in range(start_row, row_num):
                        ws.cell(row=row, column=col_idx).alignment = center_alignment
                start_row = row_num
                current_value = cell_value
        # 合并最后一批相同单元格
        if start_row < len(data) + 1:
            ws.merge_cells(
                start_row=start_row, end_row=len(data) + 1,
                start_column=col_idx, end_column=col_idx
            )
            for row in range(start_row, len(data) + 2):
                ws.cell(row=row, column=col_idx).alignment = center_alignment

    # 4. 调整列宽（自动适应内容）
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存工作簿
    # 创建res目录和时间戳文件夹（如果不存在
    timestamp = str(int(time.time() * 1000))
    folder_path = os.path.join("res_fahuo", timestamp + '_')
    os.makedirs(folder_path, exist_ok=True)
    save_path = os.path.join(folder_path, timestamp +'_发货详细表格.xlsx')
    wb.save(save_path)
    print("发货详细表格已生成：", save_path)


def Read_WareHouses_Info():
    global Path_Stores
    # warehouse_dict = defaultdict(list)
    warehouse_dict_detail = defaultdict(defaultdict)
    for warehouse in os.listdir(Path_Stores):
        warehouse_dict_sku_detail = defaultdict(list)
        # print('  ', warehouse[:4])
        # A -> SKU O ->
        warehouse_path = os.path.join(Path_Stores, warehouse)
        warehouse_4 = ''
        warehouse_pre = warehouse.split('.')[0]
        # print('----', warehouse_path)
        with open(warehouse_path, 'r', encoding='utf-8') as file:
            text = file.readlines()
            st, ed = 0, len(text)
            while st < ed and not text[st].startswith(Ware_House_Header):
            # print(warehouse, 'st:', st)
                if text[st].startswith('"货件名称"'):
                    warehouse_4 = text[st].split('","')[-1].split('-')[-1][:-2]
                st += 1 
            ######################################################
            Total_Box   = 0
            Total_Pure  = 0
            Total_V     = 0
            Total_H     = 0

            ######################################################
            # print('---- line:', st)
            for line in text[st + 1:]:
                sku_num = line.strip('\n').strip('"').split('","')
                sku_, fn_sku_, num_, box_seq = sku_num[0], sku_num[3], float(sku_num[14]), sku_num[-1]
                #TODO
                # print('sku:',sku_ , box_nums)
                Total_Box  += num_
                # print(SKU_KG[sku_][1], SKU_KG[sku_][2])
                if sku_ in SKU_KG:
                    Total_Pure += num_ * SKU_KG[sku_][1]
                    Total_V    += num_ * SKU_KG[sku_][2]
                    Total_H    += num_ * SKU_KG[sku_][2] * 167
                else:
                    print(f"Error Path_SKU_KG_Sheet '{Path_SKU_KG_Sheet}' not exits {sku_}!")
                # print('    ', sku_, ' ', num_, ' ', num_ * SKU_KG[sku_][1], ' ', num_ * SKU_KG[sku_][2], ' ', num_ * SKU_KG[sku_][2] * 167)
                box_s = box_seq.strip("'").split(",")
                box_e = str(int(box_s[0][-6:]))
                if len(box_s) > 1:
                    box_e = box_e + '-' + str(int(box_s[-1][-6:]))
                warehouse_dict_sku_detail[sku_] = [num_, fn_sku_, SKU_KG[sku_][0], box_e]
                # print(warehouse_pre + '-' + warehouse_4, warehouse_dict_sku_detail[sku_])
            # print(warehouse[:4], ',', Total_Box, ',', Total_Pure, ',', Total_V, ',', Total_H)
            # warehouse_dict[warehouse_pre] = [round(Total_Box, 2), round(Total_Pure, 2), round(Total_V, 2), round(Total_H, 2)]
            sorted_dict_sku_detail = dict(sorted(warehouse_dict_sku_detail.items(), key=lambda x: (x[1][2], x[1][3])))
            # for k, v in sorted_dict_sku_detail.items():
            #     print(warehouse_pre, v)
            warehouse_dict_detail[warehouse_pre + '-' + warehouse_4] = sorted_dict_sku_detail.copy()
    GeneratorExcelDetail(warehouse_dict_detail)

    return

def Read_Parameter():
    global Path_SKU_KG
    global Path_Stores
    # print('请输入产品预报信息表格:', end='')
    while True:
        path = input('请输入产品预报信息表格:')
        if not os.path.exists(path):
            print(f"Error: Path '{path}' not exits or file!")
            continue
        else:
            Path_SKU_KG = path
            break
    while True:
        path = input('请输入店铺最终装箱单路径:')
        if not os.path.isdir(path):
            print(f"Error: Path '{path}' not exits or dir!")
            continue
        else:
            Path_Stores = path
            break

if __name__ == '__main__':
    print('Start')
    try:
        Read_Parameter()
        Read_SKU_KG_Info()
        Read_WareHouses_Info()
    finally:
        # buffer.restore_stdout()
        pass

    print('End')
    os.system('pause')
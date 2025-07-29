import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
import time
import datetime
from collections import defaultdict
import sys
from io import StringIO

Path_HUODAIS = '/Users/lifenew/WorkSpace/Python/ExcelScrips/0603-货代报价表'
HUODAIS = defaultdict()

def Read_Parameter():
    global Path_HUODAIS
    # print('请输入产品预报信息表格:', end='')
    print('[注]货代命名格式 日期-货代 "0712-宝通达-报价单-发amz-A.xlsx"')
    while True:
        path = input('请输入货代报价表存放路径:')
        if not os.path.exists(path):
            print(f"Error: Path '{path}' not exits or file!")
            continue
        else:
            Path_HUODAIS = path
            break
    
GROUP1, GROUP2, GROUP3 = defaultdict(list), defaultdict(list), defaultdict(list)

def parse_logistics_excel(file_path):
    global GROUP1, GROUP2, GROUP3
    """
    解析物流商报价Excel表格
    
    参数:
    - file_path: Excel文件路径
    
    返回:
    - 包含所有物流信息的列表
    """
    # 打开工作簿
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    # 获取表头行（第一行）
    headers = [cell.value for cell in ws[1]]
    # print(str(headers))
    
    # 存储结果的列表
    results = []
    
    # 当前分组
    current_group = None
    
    # 从第二行开始遍历数据
    for row_idx in range(2, ws.max_row + 1):
        # 获取第一列的值（可能包含分组标记）
        first_cell_value = ws.cell(row=row_idx, column=1).value
        
        # 检查是否为分组标记
        if first_cell_value and isinstance(first_cell_value, str) and first_cell_value.startswith(('①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩')):
            current_group = first_cell_value
            # 获取当前行的数据（跳过第一列的分组标记）
            row_data = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(2, ws.max_column + 1)]
        else:
            # 普通数据行
            row_data = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        
        # 检查是否为有效数据行（至少有一个非空单元格）
        if any(row_data):
            # 创建数据字典
            data = {}
            # 如果是分组行，跳过第一列的分组标记
            if current_group and first_cell_value and first_cell_value.startswith(('①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩')):
                # 分组行的数据从第二列开始
                for col_idx, value in enumerate(row_data, 1):
                    if col_idx < len(headers):  # 确保索引不越界
                        data[headers[col_idx]] = value
            else:
                # 普通行的数据从第一列开始
                for col_idx, value in enumerate(row_data):
                    if col_idx < len(headers):  # 确保索引不越界
                        data[headers[col_idx]] = value
            
            # 添加分组信息
            if current_group:
                data['分组'] = current_group
                data['货代'] = os.path.basename(file_path).split('-')[1]
            if data['报价/KG'] == None:
                data['报价/KG'] = 99999
                # print(data['仓库'])
            if current_group and current_group == '①':
                GROUP1[data['仓库']].append(data)
            elif current_group and current_group == '②':
                GROUP2[data['仓库']].append(data)
            elif current_group and current_group == '③':
                GROUP3[data['仓库']].append(data)
            # print(str(data))
            # 添加到结果列表
            # results.append(data)
    
    # return results

def GenerateExcel():
    global GROUP1, GROUP2, GROUP3
    # 创建Workbook和工作表
    wb = Workbook()
    ws = wb.active

    # 设置标题行
    headers = [
        "时间", "店铺", "箱数", "毛重", "体积", "体积重", "地区", "仓点", "货代", 
        "报价", "渠道", "时效", "截仓时间", "船期", "下一水船期", "预计费用"
    ]
    ws.append(headers)

    # 数据内容
    data, cur_time = list(), datetime.datetime.today()
    # GROUP1 = sorted(GROUP1, key=lambda x: (x['分组'], x['仓库'], x['报价/KG']))
    for k, v_list in GROUP1.items():
        for v in v_list:
            # print(str(v))
            line = [cur_time, v['分组'], v['箱数'], v['毛重'], v['体积'], v['体积重'],
                    v['地区'], v['仓库'], v['货代'], v['报价/KG'], v['渠道'], v['时效'],
                    v['截仓时间（必填）'], v['船期'], v['下一水船期']]
            data.append(line[:])
    # GROUP2 = sorted(GROUP2, key=lambda x: (x['分组'], x['仓库'], x['报价/KG']))
    for k, v_list in GROUP2.items():
        for v in v_list:
            line = [cur_time, v['分组'], v['箱数'], v['毛重'], v['体积'], v['体积重'],
                    v['地区'], v['仓库'], v['货代'], v['报价/KG'], v['渠道'], v['时效'],
                    v['截仓时间（必填）'], v['船期'], v['下一水船期']]
            data.append(line[:])
    # GROUP3 = sorted(GROUP3, key=lambda x: (x['分组'], x['仓库'], x['报价/KG']))
    for k, v_list in GROUP3.items():
        for v in v_list:
            line = [cur_time, v['分组'], v['箱数'], v['毛重'], v['体积'], v['体积重'],
                    v['地区'], v['仓库'], v['货代'], v['报价/KG'], v['渠道'], v['时效'],
                    v['截仓时间（必填）'], v['船期'], v['下一水船期']]
            data.append(line[:])
    # 写入数据
    
    sort_data = sorted(data, key=lambda x : (x[1], x[7], float(x[9])))
    for row in sort_data:
        ws.append(row)

    # 改进后的合并函数
    # 改进的合并函数
    # 改进的合并函数
    def merge_similar_cells(ws, col_index):
        """合并指定列中相同内容的连续单元格"""
        start_row = 2  # 数据从第2行开始
        prev_value = ws.cell(row=start_row, column=col_index).value
        
        # 遍历所有行(包括最后一行之后)
        for row in range(start_row + 1, len(data) + 3):  # +3确保处理到最后一行
            try:
                current_value = ws.cell(row=row, column=col_index).value
            except:
                current_value = None  # 处理超出范围的情况
            
            # 如果值发生变化或到达最后一行
            if current_value != prev_value or row == len(data) + 2:
                # 检查是否需要合并(至少两行相同内容)
                if row - 1 > start_row and prev_value is not None and prev_value != "":
                    ws.merge_cells(
                        start_row=start_row, start_column=col_index,
                        end_row=row-1, end_column=col_index
                    )
                # 重置起始行和值
                start_row = row
                prev_value = current_value


    # 需要合并的列及其索引
    columns_to_merge = {
        "时间": 1,
        "店铺": 2,
        "箱数": 3,
        "毛重": 4,
        "体积": 5,
        "体积重": 6,
        "地区": 7,
        "仓点": 8
    }

    # 应用合并函数到每一列
    for col_name, col_idx in columns_to_merge.items():
        merge_similar_cells(ws, col_idx)

    # 设置对齐方式
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col_letter].width = adjusted_width

    # 保存文件
    # 创建res目录和时间戳文件夹（如果不存在)
    timestamp = str(int(time.time() * 1000))
    folder_path = os.path.join("res_total", str(timestamp) + '_' + 'total')
    os.makedirs(folder_path, exist_ok=True)
    save_path = os.path.join(folder_path, str(timestamp) +'_物流报价表详细表格.xlsx')
    wb.save(save_path)
    # wb.save("物流报价表_完整版.xlsx")
    # print("Excel文件已生成：物流报价表_完整版.xlsx")
    print("物流报价表详细表格已生成：", save_path)



if __name__ == '__main__':
    print('Start')
    Read_Parameter()
    for file in os.listdir(Path_HUODAIS):
        if file.endswith('.xlsx'):
            print('USE File: ', os.path.join(Path_HUODAIS, file))
            parse_logistics_excel(os.path.join(Path_HUODAIS, file))
        else:
            print(file, ' not xlsx file')
    GenerateExcel()


    print('End')
    os.system('pause')
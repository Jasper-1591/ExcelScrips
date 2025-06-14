import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

# 设置随机种子以确保结果可重现
np.random.seed(42)

# 表头（固定不变）
HEADERS = ['箱数', '毛重', '体积', '体积重', '地区', '仓库', '报价/KG', '渠道', '时效', '截仓时间（必填）', '船期', '下一水船期']
# 需要红色字体的表头
RED_HEADERS = ['箱数', '毛重', '体积', '体积重', '地区', '仓库', '截仓时间（必填）', '船期']
# 需要加粗的列
BOLD_COLUMNS = ['箱数', '仓库']

# 仓库列表
WAREHOUSES = ['FOE1', 'GSO1', 'OKC2', 'SMF6', 'XEW5', 'JVL1', 'PHX5', 'SAT4', 'ORD2']

# 生成随机数据函数
def generate_random_data(n_rows, warehouse_pool):
    """生成随机物流数据"""
    # 随机生成基础数据
    box_counts = np.random.randint(20, 150, n_rows)
    weights = np.random.uniform(200, 1500, n_rows).round(1)
    volumes = np.random.uniform(2, 12, n_rows).round(2)
    warehouses = np.random.choice(warehouse_pool, n_rows)
    
    # 计算体积重
    volume_weights = (volumes * 167).round(2)
    
    # 构建数据行
    data_rows = []
    for i in range(n_rows):
        row = {
            '箱数': box_counts[i],
            '毛重': weights[i],
            '体积': volumes[i],
            '体积重': volume_weights[i],
            '地区': '',
            '仓库': warehouses[i],
            '报价/KG': '',
            '渠道': '',
            '时效': '',
            '截仓时间（必填）': '',
            '船期': '',
            '下一水船期': ''
        }
        data_rows.append(row)
    
    return data_rows

# 生成三个部分的数据
section1_data = generate_random_data(5, ['FOE1', 'GSO1', 'OKC2', 'SMF6', 'XEW5'])
section2_data = generate_random_data(5, ['GSO1', 'JVL1', 'PHX5', 'SAT4', 'XEW5'])
section3_data = generate_random_data(5, ['ORD2', 'PHX5', 'SAT4', 'SMF6', 'XEW5'])

# 创建Excel工作簿
wb = Workbook()
ws = wb.active
ws.title = "物流数据"

# 创建字体样式
header_font = Font(bold=True)  # 表头加粗
red_font = Font(color='FF0000', bold=True)  # 红色加粗字体
bold_font = Font(bold=True)  # 加粗字体

# 写入表头，从B列开始
for col_idx, header in enumerate(HEADERS, 2):  # 从2开始，对应B列
    col_letter = get_column_letter(col_idx)
    cell = ws[f"{col_letter}1"]
    cell.value = header
    cell.alignment = Alignment(horizontal='center')  # 表头居中对齐
    
    # 设置表头字体
    if header in RED_HEADERS:
        cell.font = red_font  # 红色加粗
    else:
        cell.font = header_font  # 普通加粗

# 写入第一部分数据，从B列开始，A列合并单元格
current_row = 2
section1_end_row = current_row + len(section1_data) - 1

# 在A列合并单元格并添加标记（居中对齐）
ws.merge_cells(start_row=current_row, start_column=1, end_row=section1_end_row, end_column=1)
ws[f"A{current_row}"] = "①"
ws[f"A{current_row}"].alignment = Alignment(horizontal='center', vertical='center')  # 居中对齐

# 写入第一部分数据
for row_data in section1_data:
    for col_idx, header in enumerate(HEADERS, 2):  # 从2开始，对应B列
        col_letter = get_column_letter(col_idx)
        cell = ws[f"{col_letter}{current_row}"]
        cell.value = row_data.get(header, '')
        cell.alignment = Alignment(horizontal='left')  # 左对齐
        
        # 设置特定列字体加粗
        if header in BOLD_COLUMNS:
            cell.font = bold_font
    
    current_row += 1

# 添加第二部分
next_section_start = current_row + 1
section2_end_row = next_section_start + len(section2_data) - 1

# 在A列合并单元格并添加标记（居中对齐）
ws.merge_cells(start_row=next_section_start, start_column=1, end_row=section2_end_row, end_column=1)
ws[f"A{next_section_start}"] = "②"
ws[f"A{next_section_start}"].alignment = Alignment(horizontal='center', vertical='center')  # 居中对齐

# 写入第二部分数据
current_row = next_section_start
for row_data in section2_data:
    for col_idx, header in enumerate(HEADERS, 2):  # 从2开始，对应B列
        col_letter = get_column_letter(col_idx)
        cell = ws[f"{col_letter}{current_row}"]
        cell.value = row_data.get(header, '')
        cell.alignment = Alignment(horizontal='left')  # 左对齐
        
        # 设置特定列字体加粗
        if header in BOLD_COLUMNS:
            cell.font = bold_font
    
    current_row += 1

# 添加第三部分
next_section_start = current_row + 1
section3_end_row = next_section_start + len(section3_data) - 1

# 在A列合并单元格并添加标记（居中对齐）
ws.merge_cells(start_row=next_section_start, start_column=1, end_row=section3_end_row, end_column=1)
ws[f"A{next_section_start}"] = "③"
ws[f"A{next_section_start}"].alignment = Alignment(horizontal='center', vertical='center')  # 居中对齐

# 写入第三部分数据
current_row = next_section_start
for row_data in section3_data:
    for col_idx, header in enumerate(HEADERS, 2):  # 从2开始，对应B列
        col_letter = get_column_letter(col_idx)
        cell = ws[f"{col_letter}{current_row}"]
        cell.value = row_data.get(header, '')
        cell.alignment = Alignment(horizontal='left')  # 左对齐
        
        # 设置特定列字体加粗
        if header in BOLD_COLUMNS:
            cell.font = bold_font
    
    current_row += 1

# 自适应列宽
def auto_adjust_column_width(worksheet):
    """自动调整列宽以适应内容"""
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  # 获取列字母
        for cell in column_cells:
            try:  # 处理空单元格和非字符串值
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # 添加一些额外空间
        worksheet.column_dimensions[column].width = min(adjusted_width, 100)  # 限制最大宽度为50

# 应用列宽自适应
auto_adjust_column_width(ws)

# 保存文件
wb.save("物流数据表格.xlsx")
print("数据表格已生成：物流数据表格.xlsx")
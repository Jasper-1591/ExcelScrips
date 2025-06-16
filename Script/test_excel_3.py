import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os

GROUP1, GROUP2, GROUP3 = defaultdict(), defaultdict(), defaultdict()

def parse_logistics_excel(file_path):
    global GROUP1, GROUP2, GROUP3
    """
    解析物流商报价Excel表格
    
    参数:
    - file_path: Excel文件路径
    
    返回:
    - 包含所有物流信息的列表
    """
    # 打开工作簿
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    # 获取表头行（第一行）
    headers = [cell.value for cell in ws[1]]
    print(str(headers))
    
    # 存储结果的列表
    results = []
    
    # 当前分组
    current_group = None
    
    # 从第二行开始遍历数据
    for row_idx in range(2, ws.max_row + 1):
        # 获取第一列的值（可能包含分组标记）
        first_cell_value = ws.cell(row=row_idx, column=1).value
        
        # 检查是否为分组标记
        if first_cell_value and isinstance(first_cell_value, str) and first_cell_value.startswith(('①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩')):
            current_group = first_cell_value
            # 获取当前行的数据（跳过第一列的分组标记）
            row_data = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(2, ws.max_column + 1)]
        else:
            # 普通数据行
            row_data = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        
        # 检查是否为有效数据行（至少有一个非空单元格）
        if any(row_data):
            # 创建数据字典
            data = {}
            # 如果是分组行，跳过第一列的分组标记
            if current_group and first_cell_value and first_cell_value.startswith(('①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩')):
                # 分组行的数据从第二列开始
                for col_idx, value in enumerate(row_data, 1):
                    if col_idx < len(headers):  # 确保索引不越界
                        data[headers[col_idx]] = value
            else:
                # 普通行的数据从第一列开始
                for col_idx, value in enumerate(row_data):
                    if col_idx < len(headers):  # 确保索引不越界
                        data[headers[col_idx]] = value
            
            # 添加分组信息
            if current_group:
                data['分组'] = current_group
                data['货代'] = os.path.basename(file_path).split('-')[1]
            
            # 添加到结果列表
            results.append(data)
    
    return results

# 使用示例
if __name__ == "__main__":
    file_path = "/Users/lifenew/WorkSpace/Python/ExcelScrips/0603-货代报价表/0603-宝通达-报价单-发amz-B(2).xlsx"  # 替换为实际文件路径
    
    try:
        logistics_data = parse_logistics_excel(file_path)
        
        # 打印解析结果
        print(f"成功解析 {len(logistics_data)} 条物流信息")
        for i, data in enumerate(logistics_data, 1):  # 打印前3条记录作为示例
            print(f"\n记录 {i}:")
            for key, value in data.items():
                print(f"  {key}: {value}", end=' ')
                
    except Exception as e:
        print(f"解析过程中发生错误: {e}")
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# --------------------------
# 数据变量 - 可根据需要修改
# --------------------------

# 出货日期列表（与货件单号一一对应）
shipping_dates = [
    "0712-B店",  # 对应第一个货件单号
    "",          # 对应第二个货件单号
    "",          # 对应第三个货件单号
    ""           # 对应第四个货件单号
]

# 货件单号列表
tracking_numbers = [
    "FBA18Z3JY5M0",
    "FBA18Z3K8TMB",
    "FBA18Z3KDWCD",
    "FBA18Z3KLWQT"
]

# 物流中心编码（与货件单号一一对应）
warehouse_codes = [
    "CHA2",
    "DET1",
    "HOU8",
    "XEW5"
]

# 每个货件包含的商品数据列表
shipment_items = [
    # 第一个货件的商品
    [
        {"serial": "129-132", "sku": "NF20BP", "fnsku": "X003CAYBKB", "name": "12寸娃娃款自行车*紫色", "quantity": 4},
        {"serial": "127-128", "sku": "MARCO-12-WH", "fnsku": "X004LWHMQJ", "name": "12寸以色列男童车*白色", "quantity": 2},
        {"serial": "137-140", "sku": "Sparkle-12-GR", "fnsku": "X004PAORVZ", "name": "12寸银光款*绿色", "quantity": 4},
        {"serial": "85-87", "sku": "BixikeWF16BB", "fnsku": "X004AGCZA9", "name": "14寸带篮女童自行车*粉色", "quantity": 3},
    ],
    
    # 第二个货件的商品
    [
        {"serial": "187-190", "sku": "NF16BG", "fnsku": "X003CAPQI7", "name": "12寸娃娃款自行车*粉白(藤编车篮）", "quantity": 4},
        {"serial": "170-173", "sku": "Blossom-12-PL", "fnsku": "X004P7QBU3", "name": "12寸新款小花娃娃车*紫色", "quantity": 4},
    ],
    
    # 第三个货件的商品
    [
        {"serial": "93-100", "sku": "NF20BP", "fnsku": "X003CAYBKB", "name": "12寸娃娃款自行车*紫色", "quantity": 8},
        {"serial": "75-79", "sku": "Blossom-12-PL", "fnsku": "X004P7QBU3", "name": "12寸新款小花娃娃车*紫色", "quantity": 5},
    ],
    
    # 第四个货件的商品
    [
        {"serial": "213-222", "sku": "NF16BG", "fnsku": "X003CAPQI7", "name": "12寸娃娃款自行车*粉白(藤编车篮）", "quantity": 10},
        {"serial": "223", "sku": "NF20BP", "fnsku": "X003CAYBKB", "name": "12寸娃娃款自行车*紫色", "quantity": 1},
    ]
]

# --------------------------
# 生成基础Excel数据
# --------------------------

# 准备最终数据列表
data = {
    "出货日期": [],
    "货件单号": [],
    "Reference ID": [],
    "物流中心编码": [],
    "序列号": [],
    "SKU": [],
    "FNSKU": [],
    "品名": [],
    "发货数量": []
}

# 记录每个货件的起始和结束行索引（用于合并单元格）
shipment_row_ranges = []
current_row = 0  # 从0开始计数

# 遍历每个货件，填充数据
for i, shipment in enumerate(shipment_items):
    shipment_start_row = current_row + 1  # Excel行号从1开始
    item_count = len(shipment)
    shipment_end_row = shipment_start_row + item_count - 1
    shipment_row_ranges.append((shipment_start_row, shipment_end_row))
    
    for j, item in enumerate(shipment):
        # 只有每个货件的第一个商品行填写货件级信息，其他行留空
        if j == 0:
            data["出货日期"].append(shipping_dates[i])
            data["货件单号"].append(tracking_numbers[i])
            data["物流中心编码"].append(warehouse_codes[i])
        else:
            data["出货日期"].append("")
            data["货件单号"].append("")
            data["物流中心编码"].append("")
        
        # 商品级信息
        data["Reference ID"].append("")  # 保持为空
        data["序列号"].append(item["serial"])
        data["SKU"].append(item["sku"])
        data["FNSKU"].append(item["fnsku"])
        data["品名"].append(item["name"])
        data["发货数量"].append(item["quantity"])
        
        current_row += 1

# 创建DataFrame并导出到Excel
output_file = "带合并单元格的出货数据.xlsx"
df = pd.DataFrame(data)
df.to_excel(output_file, index=False, sheet_name="出货数据")

# --------------------------
# 处理单元格合并
# --------------------------

# 加载生成的Excel文件
wb = load_workbook(output_file)
ws = wb["出货数据"]

# 设置单元格对齐方式（垂直居中）
alignment = Alignment(vertical='center', horizontal='left')

# 合并每个货件对应的单元格
for start_row, end_row in shipment_row_ranges:
    # 合并出货日期列（A列）
    if start_row != end_row:  # 只有多行时才需要合并
        ws.merge_cells(f'A{start_row}:A{end_row}')
    # 设置对齐方式
    ws[f'A{start_row}'].alignment = alignment
    
    # 合并货件单号列（B列）
    if start_row != end_row:
        ws.merge_cells(f'B{start_row}:B{end_row}')
    ws[f'B{start_row}'].alignment = alignment
    
    # 合并物流中心编码列（D列）
    if start_row != end_row:
        ws.merge_cells(f'D{start_row}:D{end_row}')
    ws[f'D{start_row}'].alignment = alignment

# 保存修改后的Excel文件
wb.save(output_file)

print(f"已生成带合并单元格的Excel文件: {output_file}")
print(f"共包含 {len(df)} 行数据和 {len(shipment_row_ranges)} 个货件")
print(f"已自动合并相同货件的出货日期、货件单号和物流中心编码单元格")
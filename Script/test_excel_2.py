import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import random
import string


# 生成Excel文件
def GeneratorExcelDetail(output_file='刷仓数据_随机_openpyxl.xlsx'):
    """创建Excel文件并写入数据
    
    Args:
        data: 数据集
        output_file: 输出文件名
    """
    def generate_random_data(num_brushes=3, num_locations=5, num_skus=10):
        """生成随机的刷仓数据
        
        Args:
            num_brushes: 刷仓次数
            num_locations: 仓点数量
            num_skus: 每个仓点的SKU数量
        """
        data = []
        
        # 固定店铺为1个
        store = "A店"
        
        # 刷仓次数
        brushes = ["第一次刷仓", "第二次刷仓", "第三次刷仓", "第四次刷仓", "第五次刷仓"][:num_brushes]
        
        # 仓点代码
        locations = [f"{''.join(random.choices(string.ascii_uppercase, k=3))}{random.randint(1, 9)}" 
                    for _ in range(num_locations)]
        
        # 生成SKU前缀
        sku_prefixes = ["AA", "AB", "AC", "AD", "AE", "BA", "BB", "BC", "BD", "BE", 
                    "B-Doll", "Bixike", "MARCO"]
        
        # 生成数据
        for brush in brushes:
            brush_name = f"{store}_{brush}"
            
            for location in locations:
                # 每个仓点的SKU数量随机
                sku_count = random.randint(3, num_skus)
            
                for i in range(sku_count):
                    # 生成随机SKU
                    if random.random() < 0.7:  # 70%概率使用前缀
                        prefix = random.choice(sku_prefixes)
                        if prefix.startswith("B-"):
                            sku = f"{prefix}-{random.randint(10, 20)}-{random.choice(['GN', 'PK', 'WH', 'PL', 'PC'])}"
                        elif prefix.startswith("Bixike"):
                            sku = f"{prefix}{random.choice(['NF', 'WF'])}{random.randint(10, 20)}{random.choice(['BB', 'BP', 'BPK', 'BG'])}"
                        else:
                            sku = f"{prefix}{random.randint(1000, 9999)}{random.choice(['', 'FBM'])}"
                    else:  # 30%概率随机生成
                        sku = f"{''.join(random.choices(string.ascii_uppercase + string.digits, k=random.randint(5, 10)))}"
                    
                    # 生成随机数据
                    boxes = random.randint(1, 30)
                    weight = round(random.uniform(5, 15), 1)
                    total_weight = round(boxes * weight, 2)
                    volume = round(random.uniform(0.05, 0.2), 6)
                    total_volume = round(volume * boxes, 6)
                    volume_weight = round(random.uniform(10, 30), 6)
                    total_volume_weight = round(volume_weight * boxes, 6)
                    
                    # 添加到数据
                    data.append({
                        '刷仓记录': brush_name,
                        '仓点': location,
                        'SKU': sku,
                        '箱子总数': boxes,
                        '毛重': weight,
                        '总毛重': total_weight,
                        '体积': volume,
                        '总体积': total_volume,
                        '体积重': volume_weight,
                        '总体积重': total_volume_weight
                    })
        
        return data
    # 设置单元格样式
    def set_cell_style(cell, font=None, alignment=None, border=None, fill=None):
        """设置单元格样式"""
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if fill:
            cell.fill = fill

    # 写入仓点合计
    def write_location_totals(ws, headers, start_row, end_row, thin_border, total_fill):
        """为指定仓点范围写入横向合计
        
        Args:
            ws: 工作表对象
            headers: 表头列表
            start_row: 开始行
            end_row: 结束行
            thin_border: 边框样式
            total_fill: 合计单元格填充样式
        """
        sum_columns = ['箱子总数', '总毛重', '总体积', '总体积重']
        
        for header in sum_columns:
            data_col_idx = headers.index(header) + 1
            sum_col_idx = data_col_idx + 1  # 合计列在数据列的下一列
            
            # 计算合计值
            cell = ws.cell(row=start_row, column=sum_col_idx)
            
            # 构建SUM公式
            start_cell = openpyxl.utils.get_column_letter(data_col_idx) + str(start_row)
            end_cell = openpyxl.utils.get_column_letter(data_col_idx) + str(end_row)
            cell.value = f"=SUM({start_cell}:{end_cell})"
            
            # 设置样式
            set_cell_style(
                cell, 
                font=Font(bold=True, color='FF0000'),
                alignment=Alignment(horizontal='center'),
                fill=total_fill,
                border=thin_border
            )
            
            # 合并单元格
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=sum_col_idx, 
                            end_row=end_row, end_column=sum_col_idx)

    data = generate_random_data()
    print('Start')
    # 创建工作簿
    wb = openpyxl.Workbook()
    # 删除默认创建的工作表
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 创建工作表
    ws = wb.create_sheet(title='刷仓数据汇总')
    
    # 定义样式
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    total_font = Font(bold=True, color='FF0000')  # 红色字体
    location_font = Font(bold=True)
    cell_alignment = Alignment(vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    total_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')  # 黄色背景
    
    # 表头
    original_headers = ['刷仓记录', '仓点', 'SKU', '箱子总数', '毛重', '总毛重', '体积', '总体积', '体积重', '总体积重']
    headers = []
    
    # 创建包含合计列的表头
    for header in original_headers:
        headers.append(header)
        if header in ['箱子总数', '总毛重', '总体积', '总体积重']:
            headers.append(f'{header}合计')
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        set_cell_style(cell, header_font, header_alignment, thin_border)
        
        # 为合计列设置背景色
        if '合计' in header:
            cell.fill = total_fill
    
    # 如果没有数据，直接保存并返回
    if not data:
        wb.save(output_file)
        print(f"Excel文件已生成：{output_file}")
        return
    
    # 对数据按刷仓记录和仓点排序
    sorted_data = sorted(data, key=lambda x: (x['刷仓记录'], x['仓点']))
    
    # 写入数据并添加横向合计
    current_row = 2
    current_brush = None
    current_location = None
    brush_start_row = 2
    location_start_row = 2
    
    for i, row_data in enumerate(sorted_data):
        brush = row_data['刷仓记录']
        location = row_data['仓点']
        
        # 如果是新的刷仓记录，记录开始行
        if current_brush is not None and brush != current_brush:
            # 处理上一个刷仓记录的仓点合计
            location_end_row = current_row - 1
            write_location_totals(ws, headers, location_start_row, location_end_row, thin_border, total_fill)
            
            # 合并刷仓记录单元格
            ws.merge_cells(start_row=brush_start_row, start_column=1, 
                        end_row=location_end_row, end_column=1)
            
            # 记录新刷仓记录的开始行
            brush_start_row = current_row
            location_start_row = current_row
        
        # 如果是新的仓点，记录开始行
        elif current_location is not None and location != current_location:
            # 处理上一个仓点的合计
            location_end_row = current_row - 1
            write_location_totals(ws, headers, location_start_row, location_end_row, thin_border, total_fill)
            
            # 合并仓点单元格，确保与合计列一致
            ws.merge_cells(start_row=location_start_row, start_column=2, 
                        end_row=location_end_row, end_column=2)
            
            # 记录新仓点的开始行
            location_start_row = current_row
        
        # 写入当前行数据
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            
            # 只在第一行写入刷仓记录名称
            if header == '刷仓记录' and current_row == brush_start_row:
                cell.value = brush
                set_cell_style(cell, location_font, alignment=Alignment(horizontal='center', vertical='center'), border=thin_border)
            elif header == '刷仓记录':
                # 其他行刷仓记录单元格留空
                cell.value = None
                set_cell_style(cell, border=thin_border)
            
            # 只在第一行写入仓点名称
            elif header == '仓点' and current_row == location_start_row:
                cell.value = location
                set_cell_style(cell, location_font, alignment=Alignment(horizontal='center', vertical='center'), border=thin_border)
            elif header == '仓点':
                # 其他行仓点单元格留空
                cell.value = None
                set_cell_style(cell, border=thin_border)
            else:
                # 写入其他数据
                original_header = header.split('合计')[0]
                cell.value = row_data.get(original_header, '')
                set_cell_style(cell, alignment=cell_alignment, border=thin_border)
        
        current_brush = brush
        current_location = location
        current_row += 1
    
    # 处理最后一个仓点的合计
    if sorted_data:
        location_end_row = current_row - 1
        write_location_totals(ws, headers, location_start_row, location_end_row, thin_border, total_fill)
        
        # 合并最后一个仓点的单元格
        ws.merge_cells(start_row=location_start_row, start_column=2, 
                    end_row=location_end_row, end_column=2)
        
        # 处理最后一个刷仓记录的合并
        ws.merge_cells(start_row=brush_start_row, start_column=1, 
                    end_row=location_end_row, end_column=1)
    
    # 调整列宽
    for col_idx, header in enumerate(headers, 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        
        # 计算该列最大宽度
        max_length = max(
            len(str(ws.cell(row=row, column=col_idx).value)) 
            for row in range(1, current_row)
        )
        
        # 考虑标题的长度
        max_length = max(max_length, len(header)) + 2
        
        # 限制最大宽度为30
        ws.column_dimensions[column_letter].width = min(max_length, 30)
    
    # 保存工作簿
    wb.save(output_file)
    print(f"Excel文件已生成：{output_file}")


# 主函数
def main():
    # 可调整的参数
    # config = {
    #     'num_brushes': 3,         # 刷仓次数
    #     'num_locations': 5,       # 仓点数量
    #     'num_skus': 10            # 每个仓点的SKU数量
    # }
    
    # # 生成随机数据
    # data = generate_random_data(**config)
    
    # 创建Excel文件
    GeneratorExcelDetail()

    print('End')

if __name__ == "__main__":
    main()
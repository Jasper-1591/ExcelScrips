import pandas as pd
import chardet
import openpyxl
import csv
import os
import sys
import io
from collections import defaultdict

#1.  产品预报信息结构
# Excel D E F G
# sku -> [0]净重kg [1]毛重kg [2]体积m³
SKU_KG = defaultdict(list)
SKU  = 'D'
P_KG = 'E'
N_KG = 'F'
M3   = 'G'
# 产品预报信息 sku -> kg|m3|
Path_SKU_KG = '/Users/lifenew/WorkSpace/Python/ExcelTest/产品预报信息&货代询价测量表0603 - 最新的.xlsx'
Path_SKU_KG_Sheet = '产品预报明细表更新'

#2. 读取店铺刷仓记录
Paths_Stores = ['/Users/lifenew/WorkSpace/Python/ExcelTest/装箱单/A店',
                '/Users/lifenew/WorkSpace/Python/ExcelTest/装箱单/B店',
                '/Users/lifenew/WorkSpace/Python/ExcelTest/0602 B店']
Ware_House_Colums = [0, 14] # sku A
Ware_House_Header = '"SKU","商品名称"'

def Read_SKU_KG_Info():
    try:
        # 打开工作簿
        workbook = openpyxl.load_workbook(Path_SKU_KG, data_only=True)
        # 获取所有表名
        sheet_names = workbook.sheetnames

        if Path_SKU_KG_Sheet in sheet_names:
            sheet = workbook[Path_SKU_KG_Sheet]
            
            for idx in range(2, sheet.max_row + 1):
                sku_kg_data_idx = [ord(SKU) - ord('A'), ord(P_KG) - ord('A'), ord(N_KG) - ord('A'), ord(M3) - ord('A')]
                sku_kg_data = [sheet.cell(row=idx, column=i+1).value for i in sku_kg_data_idx]
                
                if all(cell is None or str(cell).strip() == '' for cell in sku_kg_data):
                    continue

                SKU_KG[sku_kg_data[0]] = [float(sku_kg_data[1]), float(sku_kg_data[2]), sku_kg_data[3]]
                # print(SKU_KG[sku_kg_data[0]])
        else:
            print(f"Error: Sheet '{Path_SKU_KG_Sheet}' not exits!")
        
    except FileNotFoundError:
        print(f"Error: File '{Path_SKU_KG}' not exits!")
    except Exception as e:
        print(f"Error: {e}")
    return None
 
def ReadStore_SumInfo():
    for storeinfo in Paths_Stores:
        try:
            if not os.path.exists(storeinfo) or not os.path.isdir(storeinfo):
                print(f"Error: Path '{storeinfo}' not exits or file!")
                return
            sorted_files = sorted(os.listdir(storeinfo))
            for store_times in sorted_files:
                store_times_path = os.path.join(storeinfo, store_times)
                if os.path.isdir(store_times_path):
                    print('store time path info: ', store_times)
                    for warehouse in os.listdir(store_times_path):
                        # print('---', warehouse[:4])
                        # A -> SKU O ->
                        warehouse_path = os.path.join(store_times_path, warehouse)
                        # print('----', warehouse_path)
                        with open(warehouse_path, 'r', encoding='utf-8') as file:
                            text = file.readlines()
                            st, ed = 0, len(text)
                            while st < ed and not text[st].startswith(Ware_House_Header):
                                # print(warehouse, 'st:', st)
                                st += 1
                            
                            ######################################################
                            Total_Box   = 0
                            Total_Pure  = 0
                            Total_V     = 0
                            Total_H     = 0

                            ######################################################
                            # print('---- line:', st)
                            for line in text[st + 1:]:
                                sku_num = line.strip('"').split('","')
                                sku_, num_ = sku_num[0], float(sku_num[14])
                                # print('----sku_:', sku_, 'num_:', num_)
                                Total_Box  += num_
                                # print(SKU_KG[sku_][1], SKU_KG[sku_][2])
                                if sku_ in SKU_KG:
                                    Total_Pure += num_ * SKU_KG[sku_][1]
                                    Total_V    += num_ * SKU_KG[sku_][2]
                                    Total_H    += num_ * SKU_KG[sku_][2] * 167
                                else:
                                    print(f"Error Path_SKU_KG_Sheet '{Path_SKU_KG_Sheet}' not exits {sku_}!")
                            print(warehouse[:4], ',', Total_Box, ',', Total_Pure, ',', Total_V, ',', Total_H)
                            ######################################################
        except Exception as e:
            print(f"Error: {e}")


if __name__ == '__main__':
    print('Start')
    Read_SKU_KG_Info()
    ReadStore_SumInfo()



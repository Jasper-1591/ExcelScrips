import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
import time
from collections import defaultdict
import sys
from io import StringIO

#1.  产品预报信息结构
# Excel D E F G
# sku -> [0]净重kg [1]毛重kg [2]体积m³
SKU_KG = defaultdict(list)
SKU  = 'B'
P_KG = 'E'
N_KG = 'F'
M3   = 'G'
# 产品预报信息 sku -> kg|m3|
Path_SKU_KG = '/Users/lifenew/WorkSpace/Python/ExcelScrips/data/0715/产品预报信息表更新0712.xlsx'
Path_SKU_KG_Sheet = '产品预报明细表更新'

#2. 读取店铺刷仓记录
Paths_Stores = ['/Users/lifenew/WorkSpace/Python/ExcelScrips/data/0715/装箱单/A']
                # '/Users/lifenew/WorkSpace/Python/ExcelTest/装箱单/B店',]
                # '/Users/lifenew/WorkSpace/Python/ExcelTest/0602 B店']
# Ware_House_Colums = [0, 14] # sku A
Ware_House_Header = '"SKU","商品名称"'

def Read_SKU_KG_Info():
    try:
        # 打开工作簿
        workbook = openpyxl.load_workbook(Path_SKU_KG, data_only=True, read_only=True)
        # 获取所有表名
        sheet_names = workbook.sheetnames

        if Path_SKU_KG_Sheet in sheet_names:
            sheet = workbook[Path_SKU_KG_Sheet]
            
            print(sheet.max_row + 1)
            for idx in range(2, sheet.max_row + 1):
                sku_kg_data_idx = [ord(SKU) - ord('A'), ord(P_KG) - ord('A'), ord(N_KG) - ord('A'), ord(M3) - ord('A')]
                sku_kg_data = [sheet.cell(row=idx, column=i+1).value for i in sku_kg_data_idx]
                
                # if sku_kg_data[0].strip() == 'Sparkle-12-GN':
                #     print('-', sku_kg_data[0].strip(), '-')
                #     print('-', sku_kg_data[0].strip(), '-')
                if all(cell is None or str(cell).strip() == '' for cell in sku_kg_data):
                    continue
                # if sku_kg_data[0].strip() == 'Sparkle-12-GN':
                #     print('-1', sku_kg_data[0].strip(), '-1')
                SKU_KG[sku_kg_data[0]] = [float(sku_kg_data[1]), float(sku_kg_data[2]), sku_kg_data[3]]
                # print(SKU_KG[sku_kg_data[0]])
        else:
            print(f"Error: Sheet '{Path_SKU_KG_Sheet}' not exits!")
        
    except FileNotFoundError:
        print(f"Error: File '{Path_SKU_KG}' not exits!")
    except Exception as e:
        print(f"Error: {e}")
    return None
 
def ReadStore_SumInfo():
    for storeinfo in Paths_Stores:
        try:
            if not os.path.exists(storeinfo) or not os.path.isdir(storeinfo):
                print(f"Error: Path '{storeinfo}' not exits or file!")
                return
            sorted_files = sorted(os.listdir(storeinfo))
            # TODO 记录三次刷仓信息
            warehouse_dict_lists = defaultdict()
            warehouse_dict_lists_detail = defaultdict()
            for store_times in sorted_files:
                store_times_path = os.path.join(storeinfo, store_times)
                if os.path.isdir(store_times_path):
                    # print('Store time path info: ', store_times)
                    warehouse_dict = defaultdict(list)
                    warehouse_dict_detail = defaultdict(defaultdict)
                    warehouse_4 = '' #warehouse[:4])
                    for warehouse in os.listdir(store_times_path):
                        warehouse_4 = '' #warehouse[:4])
                        warehouse_dict_sku_detail = defaultdict(list)
                        # print('  ', warehouse[:4])
                        # A -> SKU O ->
                        warehouse_path = os.path.join(store_times_path, warehouse)
                        # print('----', warehouse_path)
                        with open(warehouse_path, 'r', encoding='utf-8') as file:
                            text = file.readlines()
                            st, ed = 0, len(text)
                            while st < ed and not text[st].startswith(Ware_House_Header):
                                # print(warehouse, 'st:', st)
                                if text[st].startswith('"货件名称"'):
                                    warehouse_4 = text[st].split('","')[-1].split('-')[-1][:-2]
                                    # print(warehouse_4)
                                st += 1
                            
                            ######################################################
                            Total_Box   = 0
                            Total_Pure  = 0
                            Total_V     = 0
                            Total_H     = 0

                            ######################################################
                            # print('---- line:', st)
                            for line in text[st + 1:]:
                                sku_num = line.strip('"').split('","')
                                sku_, num_ = sku_num[0], float(sku_num[14])
                                # print('----sku_:', sku_, 'num_:', num_)
                                Total_Box  += num_
                                # print(SKU_KG[sku_][1], SKU_KG[sku_][2])
                                if sku_ in SKU_KG:
                                    Total_Pure += num_ * SKU_KG[sku_][1]
                                    Total_V    += num_ * SKU_KG[sku_][2]
                                    Total_H    += num_ * SKU_KG[sku_][2] * 167
                                else:
                                    print(f"Error Path_SKU_KG_Sheet '{Path_SKU_KG_Sheet}' not exits {sku_}!")
                                # print('    ', sku_, ' ', num_, ' ', num_ * SKU_KG[sku_][1], ' ', num_ * SKU_KG[sku_][2], ' ', num_ * SKU_KG[sku_][2] * 167)
                                warehouse_dict_sku_detail[sku_] = [num_, SKU_KG[sku_][1], SKU_KG[sku_][2], SKU_KG[sku_][2] * 167]
                            # print(warehouse[:4], ',', Total_Box, ',', Total_Pure, ',', Total_V, ',', Total_H)
                            if warehouse_4 == '':
                                print('warehouse_4 error:', warehouse)
                                return
                            # warehouse_dict[warehouse_4] = [round(Total_Box, 2), round(Total_Pure, 2), round(Total_V, 2), round(Total_H, 2)]
                            warehouse_dict[warehouse_4] = [Total_Box, Total_Pure, Total_V, Total_H]
                            warehouse_dict_detail[warehouse_4] = warehouse_dict_sku_detail.copy()
                            ######################################################
                    warehouse_dict_lists[store_times] = warehouse_dict.copy()
                    warehouse_dict_lists_detail[store_times] = warehouse_dict_detail.copy()
            # 获取当前时间戳（精确到毫秒）
            timestamp = str(int(time.time() * 1000))
            GeneratorExcel(warehouse_dict_lists, os.path.basename(storeinfo), timestamp)
            GeneratorExcelDetail(warehouse_dict_lists_detail, os.path.basename(storeinfo), timestamp)
        except Exception as e:
            print(f"Error: {e}")

def GeneratorExcel(warehouse_dict_lists, names, timestamp):
    global buffer
    # print(str(warehouse_dict_lists))
    # 表头（固定不变）
    HEADERS = ['箱数', '毛重', '体积', '体积重', '地区', '仓库', '报价/KG', '渠道', '时效', '截仓时间（必填）', '船期', '下一水船期']
    # 需要红色字体的表头
    RED_HEADERS = ['箱数', '毛重', '体积', '体积重', '地区', '仓库', '截仓时间（必填）', '船期']
    # 需要加粗的列
    BOLD_COLUMNS = ['箱数', '仓库']
    # 生成随机数据函数
    def generate_data(warehouse_pool):
        # 计算体积重
        # volume_weights = (volumes * 167).round(2)
        data_rows = []
        for k, v in warehouse_pool.items():
            print('k:', k, 'v:', str(v))
            row = {
                '箱数': v[0],
                '毛重': v[1],
                '体积': v[2],
                '体积重': v[3],
                '地区': '',
                '仓库': k,
                '报价/KG': '',
                '渠道': '',
                '时效': '',
                '截仓时间（必填）': '',
                '船期': '',
                '下一水船期': ''
            }
            data_rows.append(row)
            # print(str(data_rows))
        return data_rows

    generate_data_all = []
    for k, v in warehouse_dict_lists.items():
        # 生成三个部分的数据
        print('k:', k)
        generate_data_all.append(generate_data(v))

    section1_data = generate_data_all[0]
    section1_data = sorted(section1_data, key=lambda x: (x['仓库']))
    section2_data = generate_data_all[1]
    section2_data = sorted(section2_data, key=lambda x: (x['仓库']))
    section3_data = generate_data_all[2]
    section3_data = sorted(section3_data, key=lambda x: (x['仓库']))

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "物流数据"

    # 创建字体样式
    header_font = Font(bold=True)  # 表头加粗
    red_font = Font(color='FF0000', bold=True)  # 红色加粗字体
    bold_font = Font(bold=True)  # 加粗字体

    # 写入表头，从B列开始
    for col_idx, header in enumerate(HEADERS, 2):  # 从2开始，对应B列
        col_letter = get_column_letter(col_idx)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.alignment = Alignment(horizontal='center')  # 表头居中对齐
        
        # 设置表头字体
        if header in RED_HEADERS:
            cell.font = red_font  # 红色加粗
        else:
            cell.font = header_font  # 普通加粗

    # 写入第一部分数据，从B列开始，A列合并单元格
    current_row = 2
    section1_end_row = current_row + len(section1_data) - 1

    # 在A列合并单元格并添加标记（居中对齐）
    ws.merge_cells(start_row=current_row, start_column=1, end_row=section1_end_row, end_column=1)
    ws[f"A{current_row}"] = "①"
    ws[f"A{current_row}"].alignment = Alignment(horizontal='center', vertical='center')  # 居中对齐

    # 写入第一部分数据
    for row_data in section1_data:
        for col_idx, header in enumerate(HEADERS, 2):  # 从2开始，对应B列
            col_letter = get_column_letter(col_idx)
            cell = ws[f"{col_letter}{current_row}"]
            cell.value = row_data.get(header, '')
            cell.alignment = Alignment(horizontal='left')  # 左对齐
            
            # 设置特定列字体加粗
            if header in BOLD_COLUMNS:
                cell.font = bold_font
        
        current_row += 1

    # 添加第二部分
    next_section_start = current_row + 1
    section2_end_row = next_section_start + len(section2_data) - 1

    # 在A列合并单元格并添加标记（居中对齐）
    ws.merge_cells(start_row=next_section_start, start_column=1, end_row=section2_end_row, end_column=1)
    ws[f"A{next_section_start}"] = "②"
    ws[f"A{next_section_start}"].alignment = Alignment(horizontal='center', vertical='center')  # 居中对齐

    # 写入第二部分数据
    current_row = next_section_start
    for row_data in section2_data:
        for col_idx, header in enumerate(HEADERS, 2):  # 从2开始，对应B列
            col_letter = get_column_letter(col_idx)
            cell = ws[f"{col_letter}{current_row}"]
            cell.value = row_data.get(header, '')
            cell.alignment = Alignment(horizontal='left')  # 左对齐
            
            # 设置特定列字体加粗
            if header in BOLD_COLUMNS:
                cell.font = bold_font
        
        current_row += 1

    # 添加第三部分
    next_section_start = current_row + 1
    section3_end_row = next_section_start + len(section3_data) - 1

    # 在A列合并单元格并添加标记（居中对齐）
    ws.merge_cells(start_row=next_section_start, start_column=1, end_row=section3_end_row, end_column=1)
    ws[f"A{next_section_start}"] = "③"
    ws[f"A{next_section_start}"].alignment = Alignment(horizontal='center', vertical='center')  # 居中对齐

    # 写入第三部分数据
    current_row = next_section_start
    for row_data in section3_data:
        for col_idx, header in enumerate(HEADERS, 2):  # 从2开始，对应B列
            col_letter = get_column_letter(col_idx)
            cell = ws[f"{col_letter}{current_row}"]
            cell.value = row_data.get(header, '')
            cell.alignment = Alignment(horizontal='left')  # 左对齐
            
            # 设置特定列字体加粗
            if header in BOLD_COLUMNS:
                cell.font = bold_font
        
        current_row += 1

    # 自适应列宽
    def auto_adjust_column_width(worksheet):
        """自动调整列宽以适应内容"""
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter  # 获取列字母
            for cell in column_cells:
                try:  # 处理空单元格和非字符串值
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # 添加一些额外空间
            worksheet.column_dimensions[column].width = min(adjusted_width, 100)  # 限制最大宽度为50

    # 应用列宽自适应
    auto_adjust_column_width(ws)

    # 保存文件
    # 创建res目录和时间戳文件夹（如果不存在）
    folder_path = os.path.join("res", timestamp + '_' + names)
    os.makedirs(folder_path, exist_ok=True)
    # print(f"创建时间戳文件夹: {folder_path}")
    save_path = os.path.join(folder_path, timestamp +'_物流数据表格.xlsx')
    # save_path_txt = os.path.join(folder_path, timestamp +'_物流数据明细表格.txt')
    wb.save(save_path)
    # buffer.write_to_file(save_path_txt)
    print("数据表格已生成：", save_path)


def Read_Parameter():
    global Path_SKU_KG
    global Paths_Stores
    # print('请输入产品预报信息表格:', end='')
    while True:
        path = input('请输入产品预报信息表格:')
        if not os.path.exists(path):
            print(f"Error: Path '{path}' not exits or file!")
            continue
        else:
            Path_SKU_KG = path
            break
    while True:
        path = input('请输入店铺装箱表路径(存放三次记录):')
        if not os.path.isdir(path):
            print(f"Error: Path '{path}' not exits or dir!")
            continue
        else:
            Paths_Stores.clear()
            Paths_Stores.append(path)
            break
class PrintBuffer:
    """内存缓冲区，用于收集print输出"""
    def __init__(self):
        self.buffer = ""
        self.original_stdout = sys.stdout
    
    def write(self, text):
        # 收集输出到缓冲区
        self.buffer += text
    
    def flush(self):
        # 实现flush方法以满足文件对象接口
        pass
    
    def getvalue(self):
        # 获取缓冲区中的所有内容
        return self.buffer
    
    def write_to_file(self, file_path, mode='w'):
        # 将缓冲区内容写入文件
        with open(file_path, mode) as f:
            f.write(self.buffer)
    
    def restore_stdout(self):
        # 恢复标准输出
        sys.stdout = self.original_stdout

buffer = PrintBuffer()

# 生成Excel文件
def GeneratorExcelDetail(warehouse_dict_lists_detail, names, timestamp):
    """创建Excel文件并写入数据
    
    Args:
        data: 数据集
        output_file: 输出文件名
    """
    def generate_random_data(warehouse_dict_lists_detail):
        """生成随机的刷仓数据
        
        Args:
            num_brushes: 刷仓次数
            num_locations: 仓点数量
            num_skus: 每个仓点的SKU数量
        """
        data = []
        
        for brush_k, brush_v in warehouse_dict_lists_detail.items():
            # print('brush_k :', brush_k)
            for location_k, location_v in brush_v.items():
                # print('location_k :', location_k, print(str(location_v)))
                for sku_k, sku_v in location_v.items():
                    # print('sku_k: ', sku_k)
                    boxes = sku_v[0]
                    weight = sku_v[1]
                    total_weight = sku_v[1] * boxes
                    volume = sku_v[2]
                    total_volume = sku_v[2] * boxes
                    volume_weight = sku_v[3]
                    total_volume_weight = sku_v[3] * boxes
                
                    # 添加到数据
                    data.append({
                        '刷仓记录': brush_k,
                        '仓点': location_k,
                        'SKU': sku_k,
                        '箱子总数': boxes,
                        '毛重': weight,
                        '总毛重': total_weight,
                        '体积': volume,
                        '总体积': total_volume,
                        '体积重': volume_weight,
                        '总体积重': total_volume_weight
                    })
        
        return data
    # 设置单元格样式
    def set_cell_style(cell, font=None, alignment=None, border=None, fill=None):
        """设置单元格样式"""
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if fill:
            cell.fill = fill

    # 写入仓点合计
    def write_location_totals(ws, headers, start_row, end_row, thin_border, total_fill):
        """为指定仓点范围写入横向合计
        
        Args:
            ws: 工作表对象
            headers: 表头列表
            start_row: 开始行
            end_row: 结束行
            thin_border: 边框样式
            total_fill: 合计单元格填充样式
        """
        sum_columns = ['箱子总数', '总毛重', '总体积', '总体积重']
        
        for header in sum_columns:
            data_col_idx = headers.index(header) + 1
            sum_col_idx = data_col_idx + 1  # 合计列在数据列的下一列
            
            # 计算合计值
            cell = ws.cell(row=start_row, column=sum_col_idx)
            
            # 构建SUM公式
            start_cell = openpyxl.utils.get_column_letter(data_col_idx) + str(start_row)
            end_cell = openpyxl.utils.get_column_letter(data_col_idx) + str(end_row)
            cell.value = f"=SUM({start_cell}:{end_cell})"
            
            # 设置样式
            set_cell_style(
                cell, 
                font=Font(bold=True, color='FF0000'),
                alignment=Alignment(horizontal='center'),
                fill=total_fill,
                border=thin_border
            )
            
            # 合并单元格
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=sum_col_idx, 
                            end_row=end_row, end_column=sum_col_idx)

    data = generate_random_data(warehouse_dict_lists_detail)
    print('Start')
    # 创建工作簿
    wb = openpyxl.Workbook()
    # 删除默认创建的工作表
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 创建工作表
    ws = wb.create_sheet(title='刷仓数据汇总')
    
    # 定义样式
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    total_font = Font(bold=True, color='FF0000')  # 红色字体
    location_font = Font(bold=True)
    cell_alignment = Alignment(vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    total_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')  # 黄色背景
    
    # 表头
    original_headers = ['刷仓记录', '仓点', 'SKU', '箱子总数', '毛重', '总毛重', '体积', '总体积', '体积重', '总体积重']
    headers = []
    
    # 创建包含合计列的表头
    for header in original_headers:
        headers.append(header)
        if header in ['箱子总数', '总毛重', '总体积', '总体积重']:
            headers.append(f'{header}合计')
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        set_cell_style(cell, header_font, header_alignment, thin_border)
        
        # 为合计列设置背景色
        if '合计' in header:
            cell.fill = total_fill
    
    # 如果没有数据，直接保存并返回
    if not data:
        # wb.save(output_file)
        print(f"data is None{names}")
        return
    
    # 对数据按刷仓记录和仓点排序
    sorted_data = sorted(data, key=lambda x: (x['刷仓记录'], x['仓点']))
    
    # 写入数据并添加横向合计
    current_row = 2
    current_brush = None
    current_location = None
    brush_start_row = 2
    location_start_row = 2
    
    for i, row_data in enumerate(sorted_data):
        brush = row_data['刷仓记录']
        location = row_data['仓点']
        
        # 如果是新的刷仓记录，记录开始行
        if current_brush is not None and brush != current_brush:
            # 处理上一个刷仓记录的仓点合计
            location_end_row = current_row - 1
            write_location_totals(ws, headers, location_start_row, location_end_row, thin_border, total_fill)
            
            # 合并刷仓记录单元格
            ws.merge_cells(start_row=brush_start_row, start_column=1, 
                        end_row=location_end_row, end_column=1)
            
            # 记录新刷仓记录的开始行
            brush_start_row = current_row
            location_start_row = current_row
        
        # 如果是新的仓点，记录开始行
        elif current_location is not None and location != current_location:
            # 处理上一个仓点的合计
            location_end_row = current_row - 1
            write_location_totals(ws, headers, location_start_row, location_end_row, thin_border, total_fill)
            
            # 合并仓点单元格，确保与合计列一致
            ws.merge_cells(start_row=location_start_row, start_column=2, 
                        end_row=location_end_row, end_column=2)
            
            # 记录新仓点的开始行
            location_start_row = current_row
        
        # 写入当前行数据
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            
            # 只在第一行写入刷仓记录名称
            if header == '刷仓记录' and current_row == brush_start_row:
                cell.value = brush
                set_cell_style(cell, location_font, alignment=Alignment(horizontal='center', vertical='center'), border=thin_border)
            elif header == '刷仓记录':
                # 其他行刷仓记录单元格留空
                cell.value = None
                set_cell_style(cell, border=thin_border)
            
            # 只在第一行写入仓点名称
            elif header == '仓点' and current_row == location_start_row:
                cell.value = location
                set_cell_style(cell, location_font, alignment=Alignment(horizontal='center', vertical='center'), border=thin_border)
            elif header == '仓点':
                # 其他行仓点单元格留空
                cell.value = None
                set_cell_style(cell, border=thin_border)
            else:
                # 写入其他数据
                original_header = header.split('合计')[0]
                cell.value = row_data.get(original_header, '')
                set_cell_style(cell, alignment=cell_alignment, border=thin_border)
        
        current_brush = brush
        current_location = location
        current_row += 1
    
    # 处理最后一个仓点的合计
    if sorted_data:
        location_end_row = current_row - 1
        write_location_totals(ws, headers, location_start_row, location_end_row, thin_border, total_fill)
        
        # 合并最后一个仓点的单元格
        ws.merge_cells(start_row=location_start_row, start_column=2, 
                    end_row=location_end_row, end_column=2)
        
        # 处理最后一个刷仓记录的合并
        ws.merge_cells(start_row=brush_start_row, start_column=1, 
                    end_row=location_end_row, end_column=1)
    
    # 调整列宽
    for col_idx, header in enumerate(headers, 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        
        # 计算该列最大宽度
        max_length = max(
            len(str(ws.cell(row=row, column=col_idx).value)) 
            for row in range(1, current_row)
        )
        
        # 考虑标题的长度
        max_length = max(max_length, len(header)) + 2
        
        # 限制最大宽度为30
        ws.column_dimensions[column_letter].width = min(max_length, 30)
    
    # 保存工作簿
    # 创建res目录和时间戳文件夹（如果不存在）
    folder_path = os.path.join("res", timestamp + '_' + names)
    os.makedirs(folder_path, exist_ok=True)
    save_path = os.path.join(folder_path, timestamp +'_物流数据详细表格.xlsx')
    wb.save(save_path)
    print("数据详细表格已生成：", save_path)

if __name__ == '__main__':
    print('Start')
    try:
        Read_Parameter()
        # sys.stdout = buffer
        Read_SKU_KG_Info()
        ReadStore_SumInfo()
    finally:
        # buffer.restore_stdout()
        pass

    print('End')
    os.system('pause')


